# -*- coding: utf-8 -*-
"""生成产品功能清单.xlsx"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 颜色常量 ──
DEEP_BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
RED = "FF0000"
ORANGE = "FF8C00"
GRAY = "808080"

# ── 样式对象（复用） ──
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
title_font = Font(name="微软雅黑", size=16, bold=True, color=WHITE)
title_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
title_align = Alignment(horizontal="center", vertical="center")
header_font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
header_fill = PatternFill(start_color=DEEP_BLUE, end_color=DEEP_BLUE, fill_type="solid")
header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
data_font = Font(name="微软雅黑", size=10)
data_align = Alignment(vertical="center", wrap_text=True)
light_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")


def apply_title_row(ws, title, max_col):
    """第一行：合并居中标题"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = title_align
    cell.border = thin_border
    # 给合并区域其余单元格也加边框和填充
    for c in range(2, max_col + 1):
        cc = ws.cell(row=1, column=c)
        cc.fill = title_fill
        cc.border = thin_border


def apply_header_row(ws, headers, row=2):
    """第二行：列标题"""
    for idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border


def apply_data_rows(ws, data, start_row=3, priority_col=None):
    """数据行：交替颜色 + 优先级着色"""
    for r_idx, row_data in enumerate(data):
        row_num = start_row + r_idx
        fill = light_fill if r_idx % 2 == 0 else white_fill
        for c_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=value)
            cell.font = data_font
            cell.fill = fill
            cell.alignment = data_align
            cell.border = thin_border
            # 优先级着色
            if priority_col is not None and c_idx == priority_col:
                if value == "P0":
                    cell.font = Font(name="微软雅黑", size=10, bold=True, color=RED)
                elif value == "P1":
                    cell.font = Font(name="微软雅黑", size=10, bold=True, color=ORANGE)
                elif value == "P2":
                    cell.font = Font(name="微软雅黑", size=10, italic=True, color=GRAY)


def auto_col_width(ws, headers, data, min_width=10, max_width=40):
    """根据内容自适应列宽"""
    for col_idx, header in enumerate(headers, 1):
        max_len = len(str(header))
        for row in data:
            val = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
            # 粗略：中文字符算2宽度
            length = sum(2 if ord(c) > 127 else 1 for c in val)
            max_len = max(max_len, length)
        width = min(max(max_len + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ═══════════════════════════════════════════════════════════════
# Sheet 1: 业务流程分析
# ═══════════════════════════════════════════════════════════════
def build_sheet1(wb):
    headers = ["步骤编号", "流程环节", "负责角色", "输入", "处理动作", "输出", "关联系统", "痛点/价值"]
    data = [
        ["步骤1", "运营发起需求", "运营人员", "客户开票需求", "填写需求申请单提交审批", "需求审批单", "AI表格+OA审批", "在线提交替代线下填表"],
        ["步骤2", "需求汇总", "系统自动", "已审批需求", "自动汇总到需求汇总表", "需求汇总数据", "AI表格-自动化", "零人工汇总替代手工统计"],
        ["步骤3", "财务签合同", "财务人员", "已汇总需求", "创建合同并提交审批", "合同审批单", "AI表格+OA审批", "电子化签署替代线下签合同"],
        ["步骤4", "用印申请", "财务/法务", "已审批合同", "提交用印申请并审批", "用印审批单", "OA审批", "流程透明替代线下用印"],
        ["步骤5", "付款申请", "财务人员", "已用印合同", "提交付款申请并审批", "付款审批单", "AI表格+OA审批", "数据联动替代重复录入"],
        ["步骤6", "开票跟进", "运营/财务", "付款完成通知", "跟踪开票进度并确认", "开票进度记录", "AI表格-视图+自动化", "自动催办替代人工催办"],
        ["步骤7", "手续费申请", "运营人员", "开票完成记录", "提交手续费申请并审批", "手续费审批单", "OA审批", "在线审批替代线下申请"],
    ]
    ws = wb.active
    ws.title = "业务流程分析"
    apply_title_row(ws, "业务流程分析", len(headers))
    apply_header_row(ws, headers)
    apply_data_rows(ws, data)
    auto_col_width(ws, headers, data)


# ═══════════════════════════════════════════════════════════════
# Sheet 2: 解决方案架构
# ═══════════════════════════════════════════════════════════════
def build_sheet2(wb):
    ws = wb.create_sheet("解决方案架构")

    # ── 区域1：整体架构设计 ──
    h1 = ["层级", "组件", "功能说明"]
    d1 = [
        ["数据层", "钉钉AI表格", "存储所有表单数据、报表数据；支持千万热行；100+字段Agent；多视图/仪表盘/自动化"],
        ["流程层", "钉钉OA审批", "驱动业务流程流转、审批节点控制；条件分支；动态审批人；审批回调；批量审批"],
        ["智能层", "钉钉AI助手", "数据智能分析、自动提醒、预测；AI智能分析；AI仪表盘；AI翻译；AI图片识别"],
    ]

    # 区域1标题
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(h1))
    cell = ws.cell(row=1, column=1, value="整体架构设计")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = title_align
    cell.border = thin_border
    for c in range(2, len(h1) + 1):
        cc = ws.cell(row=1, column=c)
        cc.fill = title_fill
        cc.border = thin_border

    apply_header_row(ws, h1, row=2)
    apply_data_rows(ws, d1, start_row=3)
    auto_col_width(ws, h1, d1)

    # ── 区域2：核心表单设计 ──
    next_row = 3 + len(d1) + 1  # 空一行
    h2 = ["表单编号", "表单名称", "用途", "核心字段", "关联审批流程", "钉钉能力"]
    d2 = [
        ["F01", "需求申请单", "运营提交开票需求", "供应商/金额/发票类型/紧急程度", "需求审批流", "AI表格-表单"],
        ["F02", "需求汇总表", "汇总所有需求数据", "统计周期/需求笔数/总金额/完成率", "-", "AI表格-自动化+透视表"],
        ["F03", "需求任务表", "任务分配与跟踪", "任务类型/负责人/截止日期/状态", "-", "AI表格-视图"],
        ["F04", "供应商基础表", "供应商信息管理", "名称/信用代码/开户行/银行账号/联系人", "-", "AI表格-数据表"],
        ["F05", "需求匹配表", "需求与供应商匹配", "需求ID/供应商/匹配状态/匹配时间", "-", "AI表格-关联字段"],
        ["F06", "用印合同表", "合同用印申请", "合同编号/金额/签约日期/用印状态/合同附件", "用印审批流", "AI表格+OA审批"],
        ["F07", "付款申请单", "付款申请", "付款金额/付款方式/关联合同/付款状态/付款凭证", "付款审批流", "AI表格+OA审批"],
        ["F08", "开票进度表", "开票进度跟踪", "开票金额/发票号码/开票状态/申请日期/实际日期/超期天数", "-", "AI表格-视图+自动化"],
        ["F09", "手续费申请表", "手续费申请", "手续费金额/类型/关联合同/审批状态", "费用审批流", "AI表格+OA审批"],
    ]

    ws.merge_cells(start_row=next_row, start_column=1, end_row=next_row, end_column=len(h2))
    cell = ws.cell(row=next_row, column=1, value="核心表单设计")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = title_align
    cell.border = thin_border
    for c in range(2, len(h2) + 1):
        cc = ws.cell(row=next_row, column=c)
        cc.fill = title_fill
        cc.border = thin_border

    apply_header_row(ws, h2, row=next_row + 1)
    apply_data_rows(ws, d2, start_row=next_row + 2)
    # 重新计算列宽（取两个区域最大值）
    all_headers = h1 + h2
    all_data = d1 + d2
    for col_idx in range(1, max(len(h1), len(h2)) + 1):
        max_len = 0
        # 区域1
        if col_idx <= len(h1):
            max_len = max(max_len, len(str(h1[col_idx - 1])))
            for row in d1:
                val = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        # 区域2
        if col_idx <= len(h2):
            max_len = max(max_len, len(str(h2[col_idx - 1])))
            for row in d2:
                val = str(row[col_idx - 1]) if col_idx - 1 < len(row) else ""
                length = sum(2 if ord(c) > 127 else 1 for c in val)
                max_len = max(max_len, length)
        width = min(max(max_len + 2, 10), 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ═══════════════════════════════════════════════════════════════
# Sheet 3: 功能清单
# ═══════════════════════════════════════════════════════════════
def build_sheet3(wb):
    headers = ["功能编号", "一级模块", "二级模块", "功能点", "功能描述", "优先级", "AI能力应用", "钉钉能力映射", "实现方式", "验收标准"]
    data = [
        # M01 需求管理
        ["M01-01", "需求管理", "需求提交", "运营通过表单提交开票需求", "填写供应商、金额、发票类型等信息提交审批", "P0", "AI表格自动填充", "AI表格-表单", "表单+字段校验", "表单可正常提交且字段校验生效"],
        ["M01-02", "需求管理", "需求汇总", "自动汇总所有需求到汇总表", "按供应商/时间自动汇总需求笔数和金额", "P0", "AI数据分析", "AI表格-自动化", "自动化工作流", "汇总数据与明细数据一致"],
        ["M01-03", "需求管理", "需求匹配", "需求与供应商自动匹配", "根据供应商名称自动关联供应商主数据", "P1", "AI智能关联", "AI表格-关联字段", "关联字段+搜索选择", "选择供应商自动带出基础信息"],
        ["M01-04", "需求管理", "需求变更", "支持需求修改与版本追踪", "修改已提交需求并记录变更历史", "P1", "AI数据验证", "AI表格-字段", "修改记录+版本字段", "变更记录可追溯"],
        ["M01-05", "需求管理", "需求审批", "主管审核需求合理性", "部门主管审批→财务确认", "P0", "-", "OA审批", "审批流+条件分支", "审批流程正确流转"],
        # M02 合同管理
        ["M02-01", "合同管理", "合同创建", "财务创建合同关联需求", "填写合同信息并关联合同需求", "P0", "AI自动填充", "AI表格+OA审批", "表单+审批", "合同可创建且关联需求正确"],
        ["M02-02", "合同管理", "合同审批", "法务审核合同条款", "法务审核→部门负责人→财务总监", "P0", "AI文本分析", "OA审批", "审批流+条件分支", "审批流程正确流转"],
        ["M02-03", "合同管理", "合同归档", "审批通过后自动归档", "合同审批通过后自动更新状态为已签约", "P1", "-", "AI表格-自动化", "自动化工作流", "归档状态自动更新"],
        ["M02-04", "合同管理", "合同查询", "按供应商/状态/时间查询合同", "多维度筛选和搜索合同", "P1", "-", "AI表格-视图", "筛选视图+搜索", "查询结果准确"],
        # M03 用印管理
        ["M03-01", "用印管理", "用印申请", "提交用印申请关联合同", "选择合同提交用印申请", "P0", "-", "OA审批", "审批流", "申请可提交且关联合同"],
        ["M03-02", "用印管理", "用印审批", "公章管理员审批", "部门负责人→公章管理员", "P0", "-", "OA审批", "审批流", "审批流程正确"],
        ["M03-03", "用印管理", "用印状态跟踪", "实时查看用印进度", "看板视图展示各合同用印状态", "P1", "-", "AI表格-视图", "看板视图", "状态实时更新"],
        ["M03-04", "用印管理", "用印完成通知", "用印完成自动通知申请人", "用印审批通过后自动发送钉钉消息", "P1", "智能提醒", "AI表格-自动化", "自动化+消息通知", "通知及时送达"],
        # M04 付款管理
        ["M04-01", "付款管理", "付款申请", "提交付款申请关联合同", "填写付款信息并关联合同", "P0", "-", "OA审批", "审批流+条件分支", "申请可提交且关联合同"],
        ["M04-02", "付款管理", "付款审批", "财务审核→领导审批", "财务审核→条件分支(≤5万财务经理/>5万财务总监+总经理)", "P0", "-", "OA审批-条件分支", "条件审批", "条件分支正确触发"],
        ["M04-03", "付款管理", "付款执行", "审批通过后执行付款", "审批通过自动更新状态并通知出纳", "P1", "智能提醒", "AI表格-自动化", "自动化工作流", "状态自动更新且通知送达"],
        ["M04-04", "付款管理", "付款确认", "确认付款完成更新状态", "上传付款凭证并确认付款完成", "P1", "-", "AI表格", "附件字段+状态更新", "凭证可上传且状态更新"],
        # M05 开票管理
        ["M05-01", "开票管理", "开票任务生成", "合同签订后自动生成开票任务", "合同状态为已签约时自动创建开票任务", "P0", "-", "AI表格-自动化", "自动化工作流", "任务自动生成"],
        ["M05-02", "开票管理", "开票进度跟踪", "实时追踪开票状态", "看板视图展示各任务开票状态", "P0", "-", "AI表格-视图", "看板视图", "状态实时更新"],
        ["M05-03", "开票管理", "开票催办提醒", "超期未开票自动催办", "超期3天自动发送催办消息给负责人", "P1", "智能提醒", "AI表格-自动化", "定时触发+消息通知", "超期提醒准时触发"],
        ["M05-04", "开票管理", "开票确认", "确认发票已开具上传附件", "上传发票图片并确认开票完成", "P0", "AI图片识别", "AI表格", "附件字段+AI识别", "发票可上传且AI识别信息正确"],
        ["M05-05", "开票管理", "开票核验", "AI识别发票信息自动核验", "上传发票图片AI自动提取号码金额并核验", "P2", "AI图片识别", "AI字段Agent-图片识别", "AI字段Agent", "识别准确率>90%"],
        # M06 手续费管理
        ["M06-01", "手续费管理", "手续费申请", "提交手续费申请", "填写手续费金额类型关联合同", "P0", "-", "OA审批", "审批流", "申请可提交"],
        ["M06-02", "手续费管理", "手续费审批", "财务审核→领导审批", "财务审核→部门负责人→(>10万加总经理)", "P0", "-", "OA审批", "审批流+条件分支", "审批流程正确"],
        ["M06-03", "手续费管理", "手续费统计", "按供应商/时间统计手续费", "透视表按维度汇总手续费", "P1", "AI数据分析", "AI表格-透视表", "透视表视图", "统计数据准确"],
        # M07 数据分析
        ["M07-01", "数据分析", "经营看板", "全局经营数据可视化", "一屏展示需求/合同/开票/付款核心指标", "P0", "AI仪表盘", "AI表格-AI仪表盘", "AI仪表盘", "看板数据准确且实时刷新"],
        ["M07-02", "数据分析", "需求分析报表", "需求量趋势/分布分析", "按月度/季度/供应商分析需求趋势", "P1", "AI数据分析", "AI表格-透视表", "透视表视图", "趋势数据准确"],
        ["M07-03", "数据分析", "开票进度报表", "各供应商开票进度", "按供应商展示开票完成率", "P1", "-", "AI表格-仪表盘", "图表组件", "进度数据准确"],
        ["M07-04", "数据分析", "差额分析报表", "需求vs已开票差额", "自动计算需求金额与已开票金额差值", "P1", "AI智能公式", "AI表格-公式", "智能公式", "差额计算正确"],
        ["M07-05", "数据分析", "费用汇总报表", "手续费/付款汇总", "按月度/类型汇总费用", "P1", "AI数据分析", "AI表格-透视表", "透视表视图", "汇总数据准确"],
        ["M07-06", "数据分析", "AI智能分析", "对话式数据问答", "通过AI助理对话查询业务数据", "P2", "AI智能分析", "AI表格-AI助理", "AI助理", "问答结果准确"],
        # M08 系统管理
        ["M08-01", "系统管理", "供应商管理", "供应商基础信息维护", "增删改查供应商信息", "P0", "AI自动填充", "AI表格", "数据表+表单", "供应商信息可正常维护"],
        ["M08-02", "系统管理", "权限管理", "按角色配置数据/功能权限", "按角色设置数据查看/编辑权限", "P0", "-", "AI表格-高级权限", "角色权限", "权限控制生效"],
        ["M08-03", "系统管理", "流程配置", "审批流程节点/条件配置", "配置审批节点审批人条件分支超时策略", "P0", "-", "OA审批", "流程设计", "流程配置正确"],
        ["M08-04", "系统管理", "数据备份", "定期数据备份与恢复", "导出数据表为Excel备份", "P2", "-", "AI表格", "导出功能", "备份文件完整"],
    ]

    ws = wb.create_sheet("功能清单")
    apply_title_row(ws, "功能清单", len(headers))
    apply_header_row(ws, headers)
    apply_data_rows(ws, data, priority_col=6)
    auto_col_width(ws, headers, data, max_width=50)
    # 冻结窗格：冻结前3行（标题+列标题）和第1列
    ws.freeze_panes = "B3"


# ═══════════════════════════════════════════════════════════════
# Sheet 4: AI智能功能
# ═══════════════════════════════════════════════════════════════
def build_sheet4(wb):
    headers = ["编号", "功能点", "说明", "AI能力应用", "价值", "实现方式"]
    data = [
        ["AI-01", "智能提醒", "逾期未处理自动提醒", "AI表格自动化+消息通知", "提升效率避免遗漏", "配置定时触发规则"],
        ["AI-02", "开票预测", "根据历史数据预测开票周期", "AI智能分析", "辅助决策合理安排资源", "AI助理对话式分析"],
        ["AI-03", "异常检测", "自动识别数据异常", "AI智能公式+条件格式", "风险预警提前发现问题", "公式计算+标红高亮"],
        ["AI-04", "报表生成", "自动生成各类统计报表", "AI仪表盘+透视表", "节省人力一键生成报表", "AI仪表盘自动生成"],
        ["AI-05", "发票识别", "上传发票图片自动提取信息", "AI字段Agent-图片识别", "减少80%手工录入", "AI字段自动识别"],
        ["AI-06", "智能填报", "输入关键词自动补全信息", "AI字段Agent-AI生成文本", "提升填报效率60%", "AI字段自动生成"],
        ["AI-07", "数据同步", "审批数据自动同步到表格", "数据连接中心", "零人工数据搬运", "OA审批数据同步配置"],
        ["AI-08", "智能翻译", "供应商信息自动翻译", "AI字段Agent-AI翻译", "支持跨国业务13种语言", "AI翻译字段"],
    ]
    ws = wb.create_sheet("AI智能功能")
    apply_title_row(ws, "AI智能功能", len(headers))
    apply_header_row(ws, headers)
    apply_data_rows(ws, data)
    auto_col_width(ws, headers, data)


# ═══════════════════════════════════════════════════════════════
# Sheet 5: 报表功能
# ═══════════════════════════════════════════════════════════════
def build_sheet5(wb):
    headers = ["报表编号", "报表名称", "统计维度", "用途", "数据源", "核心指标", "钉钉能力", "刷新频率"]
    data = [
        ["R-01", "总需求量报表", "按时间/部门/供应商统计发票金额", "预算规划", "需求申请表", "需求笔数/总金额/同比增长", "AI表格-透视表", "每日"],
        ["R-02", "已开票金额报表", "统计已完成开票金额", "财务对账", "开票进度表", "已开票笔数/已开票金额/完成率", "AI表格-透视表", "每日"],
        ["R-03", "差额分析报表", "需求金额与已开票金额差值", "进度追踪", "需求申请表+开票进度表", "差额/差额占比/趋势", "AI表格-智能公式", "实时"],
        ["R-04", "开票进度报表", "各供应商开票进度百分比", "进度监控", "开票进度表", "各供应商完成率/超期数/平均开票天数", "AI表格-仪表盘", "实时"],
        ["R-05", "手续费汇总报表", "统计手续费总额", "成本核算", "手续费申请表", "手续费总额/占比/趋势", "AI表格-透视表", "每日"],
        ["R-06", "经营总览看板", "全局核心指标一屏展示", "经营决策", "全部数据表", "需求总额/已签约额/已开票额/差额/完成率", "AI表格-AI仪表盘", "实时"],
        ["R-07", "审批效率报表", "审批流程效率分析", "流程优化", "OA审批同步数据", "平均审批时长/超时率/瓶颈节点", "AI表格-透视表", "每周"],
        ["R-08", "超期预警看板", "超期未处理事项预警", "风险管控", "开票进度表+付款申请表", "超期笔数/最长超期天数/责任人", "AI表格-看板视图+自动化", "实时"],
    ]
    ws = wb.create_sheet("报表功能")
    apply_title_row(ws, "报表功能", len(headers))
    apply_header_row(ws, headers)
    apply_data_rows(ws, data)
    auto_col_width(ws, headers, data)


# ═══════════════════════════════════════════════════════════════
# 主函数
# ═══════════════════════════════════════════════════════════════
def main():
    wb = Workbook()
    build_sheet1(wb)
    build_sheet2(wb)
    build_sheet3(wb)
    build_sheet4(wb)
    build_sheet5(wb)

    output = "产品功能清单.xlsx"
    wb.save(output)
    print(f"[OK] 已生成: {output}")


if __name__ == "__main__":
    main()
