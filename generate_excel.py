#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
钉钉AI表格+OA审批解决方案 Excel生成脚本
运行后生成: 钉钉AI表格+OA审批解决方案.xlsx
"""

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, NamedStyle
)
from openpyxl.utils import get_column_letter


# ── 全局样式定义 ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(name="微软雅黑", bold=True, size=16, color="1F4E79")
SUBTITLE_FONT = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")
NORMAL_FONT = Font(name="微软雅黑", size=10)
BOLD_FONT = Font(name="微软雅黑", bold=True, size=10)
ROW_FILL_EVEN = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
ROW_FILL_ODD = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)
WRAP_ALIGN = Alignment(horizontal="left", vertical="top", wrap_text=True)

# 优先级颜色
P0_FONT = Font(name="微软雅黑", size=10, bold=True, color="FF0000")
P1_FONT = Font(name="微软雅黑", size=10, bold=True, color="ED7D31")
P2_FONT = Font(name="微软雅黑", size=10, color="808080")

# 风险等级背景
RISK_HIGH_FILL = PatternFill(start_color="FF4444", end_color="FF4444", fill_type="solid")
RISK_HIGH_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
RISK_MID_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
RISK_MID_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
RISK_LOW_FILL = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
RISK_LOW_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")


def apply_title(ws, title, max_col):
    """在A1写入合并居中的Sheet标题"""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    ws.row_dimensions[1].height = 40


def apply_header(ws, row, headers):
    """写入表头行"""
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 28


def apply_data_row(ws, row, values, row_idx=0, alignments=None):
    """写入数据行，交替背景色"""
    fill = ROW_FILL_EVEN if row_idx % 2 == 0 else ROW_FILL_ODD
    for col_idx, val in enumerate(values, 1):
        cell = ws.cell(row=row, column=col_idx, value=val)
        cell.font = NORMAL_FONT
        cell.fill = fill
        cell.border = THIN_BORDER
        if alignments and col_idx <= len(alignments):
            cell.alignment = alignments[col_idx - 1]
        else:
            cell.alignment = LEFT_ALIGN


def auto_column_width(ws, min_width=10, max_width=45):
    """自适应列宽"""
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                # 粗略估算：中文占2个字符宽
                length = sum(2 if ord(c) > 127 else 1 for c in val_str)
                if length > max_len:
                    max_len = length
        adjusted = min(max(max_len + 4, min_width), max_width)
        ws.column_dimensions[col_letter].width = adjusted


def freeze_first_col(ws):
    """冻结第一列"""
    ws.freeze_panes = "B1"


def apply_priority_font(ws, row, col):
    """根据优先级值设置字体颜色"""
    cell = ws.cell(row=row, column=col)
    val = str(cell.value).strip() if cell.value else ""
    if val == "P0":
        cell.font = P0_FONT
    elif val == "P1":
        cell.font = P1_FONT
    elif val == "P2":
        cell.font = P2_FONT


def apply_risk_level(ws, row, col):
    """根据风险等级设置背景色"""
    cell = ws.cell(row=row, column=col)
    val = str(cell.value).strip() if cell.value else ""
    if val == "高":
        cell.fill = RISK_HIGH_FILL
        cell.font = RISK_HIGH_FONT
    elif val == "中":
        cell.fill = RISK_MID_FILL
        cell.font = RISK_MID_FONT
    elif val == "低":
        cell.fill = RISK_LOW_FILL
        cell.font = RISK_LOW_FONT


# ══════════════════════════════════════════════════════════════
# Sheet 1: 项目概览
# ══════════════════════════════════════════════════════════════
def create_sheet_01(wb):
    ws = wb.active
    ws.title = "项目概览"

    max_col = 2
    apply_title(ws, "钉钉AI表格+OA审批 开票流程数字化解决方案", max_col)

    # 基本信息表
    info_headers = ["项目", "内容"]
    apply_header(ws, 3, info_headers)

    info_data = [
        ["项目名称", "钉钉AI表格+OA审批 开票流程数字化解决方案"],
        ["文档版本", "V2.0"],
        ["编制日期", "2026-06-06"],
        ["文档性质", "解决方案与功能清单"],
    ]
    for i, row_data in enumerate(info_data):
        apply_data_row(ws, 4 + i, row_data, i)
        ws.cell(row=4 + i, column=1).font = BOLD_FONT

    # 项目背景
    r = 9
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="项目背景")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cell.border = THIN_BORDER

    r = 10
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1,
                   value="企业开票流程涉及运营、财务、法务多部门协作，当前依赖线下表格和手工流转，存在数据孤岛、流程不透明、效率低下等痛点。")
    cell.font = NORMAL_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER
    ws.row_dimensions[r].height = 36

    # 项目目标
    r = 12
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="项目目标")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cell.border = THIN_BORDER

    r = 13
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1,
                   value='基于钉钉AI表格+OA审批能力，构建"数据驱动+流程自动化+AI赋能"的开票流程数字化体系，实现全流程在线化、自动化、智能化。')
    cell.font = NORMAL_FONT
    cell.alignment = WRAP_ALIGN
    cell.border = THIN_BORDER
    ws.row_dimensions[r].height = 36

    # 核心痛点
    r = 15
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="核心痛点")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cell.border = THIN_BORDER

    pain_headers = ["序号", "痛点描述"]
    apply_header(ws, 16, pain_headers)

    pains = [
        ["1", "数据孤岛：各环节使用独立Excel，数据无法实时共享"],
        ["2", "流程割裂：审批与数据管理分离，需二次录入"],
        ["3", "进度不透明：无法实时追踪开票进度，依赖人工催办"],
        ["4", "统计滞后：报表依赖人工汇总，决策数据延迟"],
        ["5", "风险不可控：缺乏异常预警机制，超期/超额风险难发现"],
    ]
    for i, row_data in enumerate(pains):
        apply_data_row(ws, 17 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 2: 业务流程分析
# ══════════════════════════════════════════════════════════════
def create_sheet_02(wb):
    ws = wb.create_sheet("业务流程分析")
    max_col = 4
    apply_title(ws, "业务流程分析 — AS-IS vs TO-BE", max_col)

    # AS-IS
    r = 3
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="AS-IS 现状流程")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    cell.border = THIN_BORDER

    headers = ["步骤", "流程环节", "痛点", "影响"]
    apply_header(ws, 4, headers)

    as_is = [
        ["1", "运营线下填表提交开票需求", "纸质/Excel单据易丢失", "数据遗漏风险高"],
        ["2", "需求人工汇总到总表", "手工汇总耗时易错", "效率低、错误率高"],
        ["3", "财务线下签合同", "合同版本混乱", "法律风险"],
        ["4", "线下提交用印申请", "用印状态无法追踪", "流程不透明"],
        ["5", "线下提交付款申请", "付款进度不透明", "资金管控难"],
        ["6", "人工跟踪开票进度", "进度依赖人工催办", "响应滞后"],
    ]
    for i, row_data in enumerate(as_is):
        apply_data_row(ws, 5 + i, row_data, i)

    # TO-BE
    r = 12
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="TO-BE 目标流程")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    cell.border = THIN_BORDER

    apply_header(ws, 13, headers)

    to_be = [
        ["1", "运营通过AI表格表单提交需求", "在线提交，数据结构化", "零纸质、零遗漏"],
        ["2", "需求自动汇总到AI表格", "实时汇总，零人工", "效率提升90%"],
        ["3", "财务在线签合同，OA审批流转", "电子化签署，全程留痕", "合规可追溯"],
        ["4", "用印申请OA审批，状态实时可查", "流程透明，自动提醒", "进度实时可见"],
        ["5", "付款申请OA审批，关联合同数据", "数据联动，防错防漏", "准确率提升"],
        ["6", "开票进度AI表格自动跟踪", "自动催办，智能预警", "超期率降低80%"],
    ]
    for i, row_data in enumerate(to_be):
        apply_data_row(ws, 14 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 3: 解决方案架构
# ══════════════════════════════════════════════════════════════
def create_sheet_03(wb):
    ws = wb.create_sheet("解决方案架构")
    max_col = 3
    apply_title(ws, "解决方案架构设计", max_col)

    headers = ["架构层级", "核心组件", "说明"]
    apply_header(ws, 3, headers)

    arch_data = [
        ["展示层", "钉钉工作台 | 消息通知 | 待办中心 | 仪表盘", "用户统一入口，多端适配"],
        ["应用层", "需求管理 | 合同管理 | 用印管理 | 付款管理 | 开票管理 | 手续费管理", "六大业务模块覆盖全流程"],
        ["能力层", "AI表格(数据表/表单/视图/仪表盘/自动化/高级权限)", "数据管理核心能力"],
        ["能力层", "OA审批(流程引擎/条件分支/动态审批人/审批回调/批量审批)", "流程自动化核心能力"],
        ["能力层", "AI能力(字段Agent/智能分析/AI仪表盘/透视表/智能公式)", "智能化增强能力"],
        ["数据层", "供应商主数据 | 需求数据 | 合同数据 | 付款数据 | 开票数据 | 审批数据", "结构化数据资产"],
    ]
    for i, row_data in enumerate(arch_data):
        apply_data_row(ws, 4 + i, row_data, i)

    # 技术架构说明
    r = 11
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="技术架构说明")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cell.border = THIN_BORDER

    tech_headers = ["技术维度", "技术选型", "说明"]
    apply_header(ws, 12, tech_headers)

    tech_data = [
        ["前端", "钉钉原生客户端（PC/移动端）", "零开发，开箱即用"],
        ["后端", "钉钉云+企业自有系统", "混合部署，灵活扩展"],
        ["数据存储", "AI表格（千万热行级）+ 数据连接中心", "海量数据高性能存储"],
        ["集成", "开放API + 连接器 + Stream推送", "与企业现有系统无缝对接"],
    ]
    for i, row_data in enumerate(tech_data):
        apply_data_row(ws, 13 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 4: 功能模块清单
# ══════════════════════════════════════════════════════════════
def create_sheet_04(wb):
    ws = wb.create_sheet("功能模块清单")
    max_col = 8
    apply_title(ws, "功能模块清单", max_col)

    headers = ["模块编号", "一级模块", "二级模块", "功能点", "功能描述", "优先级", "钉钉能力映射", "实现方式"]
    apply_header(ws, 3, headers)

    modules = [
        ["M01-01", "需求管理", "需求提交", "运营通过表单提交开票需求", "运营通过表单提交开票需求（供应商、金额、发票类型等）", "P0", "AI表格-表单", "表单+自动化"],
        ["M01-02", "需求管理", "需求汇总", "自动汇总所有需求到汇总表", "自动汇总所有需求到汇总表", "P0", "AI表格-自动化", "自动化工作流"],
        ["M01-03", "需求管理", "需求匹配", "需求与供应商自动匹配", "需求与供应商自动匹配", "P1", "AI表格-关联字段", "关联+公式"],
        ["M01-04", "需求管理", "需求变更", "支持需求修改与版本追踪", "支持需求修改与版本追踪", "P1", "AI表格-字段", "修改记录"],
        ["M01-05", "需求管理", "需求审批", "主管审核需求合理性", "主管审核需求合理性", "P0", "OA审批", "审批流"],

        ["M02-01", "合同管理", "合同创建", "财务创建合同，关联需求", "财务创建合同，关联需求", "P0", "AI表格+OA审批", "表单+审批"],
        ["M02-02", "合同管理", "合同审批", "法务审核合同条款", "法务审核合同条款", "P0", "OA审批", "审批流"],
        ["M02-03", "合同管理", "合同归档", "审批通过后自动归档", "审批通过后自动归档", "P1", "AI表格-自动化", "自动化工作流"],
        ["M02-04", "合同管理", "合同查询", "按供应商/状态/时间查询", "按供应商/状态/时间查询", "P1", "AI表格-视图", "筛选视图"],

        ["M03-01", "用印管理", "用印申请", "提交用印申请，关联合同", "提交用印申请，关联合同", "P0", "OA审批", "审批流"],
        ["M03-02", "用印管理", "用印审批", "公章管理员审批", "公章管理员审批", "P0", "OA审批", "审批流"],
        ["M03-03", "用印管理", "用印状态跟踪", "实时查看用印进度", "实时查看用印进度", "P1", "AI表格-视图", "看板视图"],
        ["M03-04", "用印管理", "用印完成通知", "用印完成自动通知", "用印完成自动通知", "P1", "AI表格-自动化", "消息通知"],

        ["M04-01", "付款管理", "付款申请", "提交付款申请，关联合同", "提交付款申请，关联合同", "P0", "OA审批", "审批流"],
        ["M04-02", "付款管理", "付款审批", "财务审核→领导审批", "财务审核→领导审批", "P0", "OA审批-条件分支", "条件审批"],
        ["M04-03", "付款管理", "付款执行", "审批通过后执行付款", "审批通过后执行付款", "P1", "AI表格-自动化", "自动化工作流"],
        ["M04-04", "付款管理", "付款确认", "确认付款完成，更新状态", "确认付款完成，更新状态", "P1", "AI表格", "状态更新"],

        ["M05-01", "开票管理", "开票任务生成", "合同签订后自动生成开票任务", "合同签订后自动生成开票任务", "P0", "AI表格-自动化", "自动化工作流"],
        ["M05-02", "开票管理", "开票进度跟踪", "实时追踪开票状态", "实时追踪开票状态", "P0", "AI表格-视图", "看板视图"],
        ["M05-03", "开票管理", "开票催办提醒", "超期未开票自动催办", "超期未开票自动催办", "P1", "AI表格-自动化", "定时触发"],
        ["M05-04", "开票管理", "开票确认", "确认发票已开具，上传附件", "确认发票已开具，上传附件", "P0", "AI表格", "附件字段"],
        ["M05-05", "开票管理", "开票核验", "AI识别发票信息自动核验", "AI识别发票信息自动核验", "P2", "AI表格-AI字段Agent", "图片识别"],

        ["M06-01", "手续费管理", "手续费申请", "提交手续费申请", "提交手续费申请", "P0", "OA审批", "审批流"],
        ["M06-02", "手续费管理", "手续费审批", "财务审核→领导审批", "财务审核→领导审批", "P0", "OA审批", "审批流"],
        ["M06-03", "手续费管理", "手续费统计", "按供应商/时间统计", "按供应商/时间统计", "P1", "AI表格-透视表", "透视表视图"],

        ["M07-01", "数据分析", "经营看板", "全局经营数据可视化", "全局经营数据可视化", "P0", "AI表格-AI仪表盘", "AI仪表盘"],
        ["M07-02", "数据分析", "需求分析报表", "需求量趋势/分布分析", "需求量趋势/分布分析", "P1", "AI表格-透视表", "透视表视图"],
        ["M07-03", "数据分析", "开票进度报表", "各供应商开票进度", "各供应商开票进度", "P1", "AI表格-仪表盘", "图表组件"],
        ["M07-04", "数据分析", "差额分析报表", "需求vs已开票差额", "需求vs已开票差额", "P1", "AI表格-公式", "智能公式"],
        ["M07-05", "数据分析", "费用汇总报表", "手续费/付款汇总", "手续费/付款汇总", "P1", "AI表格-透视表", "透视表视图"],
        ["M07-06", "数据分析", "AI智能分析", "对话式数据问答", "对话式数据问答", "P2", "AI表格-AI助理", "AI智能分析"],

        ["M08-01", "系统管理", "供应商管理", "供应商基础信息维护", "供应商基础信息维护", "P0", "AI表格", "数据表"],
        ["M08-02", "系统管理", "权限管理", "按角色配置数据/功能权限", "按角色配置数据/功能权限", "P0", "AI表格-高级权限", "角色权限"],
        ["M08-03", "系统管理", "流程配置", "审批流程节点/条件配置", "审批流程节点/条件配置", "P0", "OA审批", "流程设计"],
        ["M08-04", "系统管理", "数据备份", "定期数据备份与恢复", "定期数据备份与恢复", "P2", "AI表格", "导出功能"],
    ]

    for i, row_data in enumerate(modules):
        apply_data_row(ws, 4 + i, row_data, i)
        # 优先级着色
        apply_priority_font(ws, 4 + i, 6)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 5: 数据模型设计
# ══════════════════════════════════════════════════════════════
def create_sheet_05(wb):
    ws = wb.create_sheet("数据模型设计")
    max_col = 5
    apply_title(ws, "数据模型设计", max_col)

    headers = ["数据表", "字段名称", "字段类型", "约束/说明", "备注"]
    apply_header(ws, 3, headers)

    tables = [
        # 供应商主数据表
        ["供应商主数据表", "供应商ID", "文本", "主键", "自动生成"],
        ["供应商主数据表", "供应商名称", "文本", "必填", ""],
        ["供应商主数据表", "统一社会信用代码", "文本", "唯一", "18位"],
        ["供应商主数据表", "开户银行", "文本", "", ""],
        ["供应商主数据表", "银行账号", "文本", "", ""],
        ["供应商主数据表", "联系人", "文本", "", ""],
        ["供应商主数据表", "联系电话", "文本", "", ""],
        ["供应商主数据表", "合作状态", "单选", "选项:合作中/暂停/终止", ""],
        ["供应商主数据表", "首次合作日期", "日期", "", ""],
        ["供应商主数据表", "备注", "文本", "", ""],

        # 需求申请表
        ["需求申请表", "需求ID", "文本", "主键", "自动生成"],
        ["需求申请表", "申请日期", "日期", "必填", ""],
        ["需求申请表", "申请人", "人员", "必填", "钉钉人员字段"],
        ["需求申请表", "所属部门", "部门", "必填", "钉钉部门字段"],
        ["需求申请表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["需求申请表", "发票类型", "单选", "选项:专票/普票", ""],
        ["需求申请表", "发票金额", "数字", "必填", "不含税"],
        ["需求申请表", "税额", "数字", "公式", "=发票金额×税率"],
        ["需求申请表", "价税合计", "数字", "公式", "=发票金额+税额"],
        ["需求申请表", "需求说明", "文本", "", ""],
        ["需求申请表", "紧急程度", "单选", "选项:普通/紧急", ""],
        ["需求申请表", "审批状态", "单选", "选项:待审批/审批中/已通过/已拒绝", ""],
        ["需求申请表", "关联合同ID", "关联字段", "→合同表", "关联查询"],
        ["需求申请表", "创建时间", "日期", "自动", "系统生成"],
        ["需求申请表", "更新时间", "日期", "自动", "系统生成"],

        # 需求汇总表
        ["需求汇总表", "汇总ID", "文本", "主键", "自动生成"],
        ["需求汇总表", "统计周期", "单选", "选项:月/季/年", ""],
        ["需求汇总表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["需求汇总表", "需求笔数", "数字", "公式COUNT", "自动统计"],
        ["需求汇总表", "需求总金额", "数字", "公式SUM", "自动汇总"],
        ["需求汇总表", "已签约金额", "数字", "公式SUM", "自动汇总"],
        ["需求汇总表", "差额", "数字", "公式", "=需求总金额-已签约金额"],
        ["需求汇总表", "完成率", "数字", "公式", "=已签约金额/需求总金额"],

        # 合同管理表
        ["合同管理表", "合同ID", "文本", "主键", "自动生成"],
        ["合同管理表", "合同编号", "文本", "唯一", ""],
        ["合同管理表", "合同名称", "文本", "必填", ""],
        ["合同管理表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["合同管理表", "合同金额", "数字", "必填", ""],
        ["合同管理表", "签约日期", "日期", "", ""],
        ["合同管理表", "合同状态", "单选", "选项:草稿/审批中/已签约/已终止", ""],
        ["合同管理表", "关联需求ID", "关联字段", "→需求申请表", "关联查询"],
        ["合同管理表", "合同附件", "附件", "", "支持PDF/图片"],
        ["合同管理表", "用印状态", "单选", "选项:待用印/用印中/已完成", ""],
        ["合同管理表", "审批状态", "单选", "选项:待审批/审批中/已通过/已拒绝", ""],
        ["合同管理表", "创建人", "人员", "自动", "钉钉人员字段"],
        ["合同管理表", "创建时间", "日期", "自动", "系统生成"],

        # 付款申请表
        ["付款申请表", "付款ID", "文本", "主键", "自动生成"],
        ["付款申请表", "申请日期", "日期", "必填", ""],
        ["付款申请表", "申请人", "人员", "必填", "钉钉人员字段"],
        ["付款申请表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["付款申请表", "付款金额", "数字", "必填", ""],
        ["付款申请表", "付款方式", "单选", "选项:银行转账/承兑汇票", ""],
        ["付款申请表", "关联合同ID", "关联字段", "→合同管理表", "关联查询"],
        ["付款申请表", "付款状态", "单选", "选项:待审批/审批中/已付款/已拒绝", ""],
        ["付款申请表", "审批状态", "单选", "选项:待审批/审批中/已通过/已拒绝", ""],
        ["付款申请表", "付款凭证", "附件", "", "支持PDF/图片"],
        ["付款申请表", "付款日期", "日期", "", ""],

        # 开票进度表
        ["开票进度表", "开票ID", "文本", "主键", "自动生成"],
        ["开票进度表", "关联需求ID", "关联字段", "→需求申请表", "关联查询"],
        ["开票进度表", "关联合同ID", "关联字段", "→合同管理表", "关联查询"],
        ["开票进度表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["开票进度表", "开票金额", "数字", "必填", ""],
        ["开票进度表", "发票号码", "文本", "", ""],
        ["开票进度表", "发票类型", "单选", "选项:专票/普票", ""],
        ["开票进度表", "开票状态", "单选", "选项:待开票/开票中/已开票/已作废", ""],
        ["开票进度表", "申请开票日期", "日期", "", ""],
        ["开票进度表", "实际开票日期", "日期", "", ""],
        ["开票进度表", "发票附件", "附件/图片", "", "支持图片上传"],
        ["开票进度表", "AI识别结果", "AI字段-图片识别", "", "AI自动提取"],
        ["开票进度表", "超期天数", "数字", "公式", "=TODAY()-申请开票日期"],
        ["开票进度表", "催办次数", "数字", "", "默认0"],

        # 手续费申请表
        ["手续费申请表", "手续费ID", "文本", "主键", "自动生成"],
        ["手续费申请表", "申请日期", "日期", "必填", ""],
        ["手续费申请表", "申请人", "人员", "必填", "钉钉人员字段"],
        ["手续费申请表", "供应商", "关联字段", "→供应商主数据", "关联查询"],
        ["手续费申请表", "手续费金额", "数字", "必填", ""],
        ["手续费申请表", "手续费类型", "单选", "选项:服务费/代理费/其他", ""],
        ["手续费申请表", "关联合同ID", "关联字段", "→合同管理表", "关联查询"],
        ["手续费申请表", "审批状态", "单选", "选项:待审批/审批中/已通过/已拒绝", ""],
        ["手续费申请表", "备注", "文本", "", ""],
    ]

    for i, row_data in enumerate(tables):
        apply_data_row(ws, 4 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 6: 审批流程设计
# ══════════════════════════════════════════════════════════════
def create_sheet_06(wb):
    ws = wb.create_sheet("审批流程设计")
    max_col = 5
    apply_title(ws, "审批流程设计", max_col)

    headers = ["流程名称", "流程节点", "条件分支", "超时策略", "备注"]
    apply_header(ws, 3, headers)

    flows = [
        ["需求审批流",
         "发起人提交 → 部门主管审批 → 财务部确认 → 结束",
         "金额>50万增加总经理审批节点",
         "24小时未处理自动提醒，72小时升级",
         ""],
        ["合同审批流",
         "财务发起 → 法务审核 → 部门负责人审批 → 财务总监审批 → 结束",
         "金额>100万增加CEO审批",
         "24小时未处理自动提醒",
         ""],
        ["用印审批流",
         "申请人提交 → 部门负责人审批 → 公章管理员审批 → 结束",
         "合同类型=特殊合同增加法务审核",
         "12小时未处理自动提醒",
         ""],
        ["付款审批流",
         "申请人提交 → 财务审核 → 条件分支(金额≤5万→财务经理审批 / 金额>5万→财务总监审批→总经理审批) → 结束",
         "金额≤5万→财务经理审批；金额>5万→财务总监审批→总经理审批",
         "24小时未处理自动提醒，48小时升级",
         ""],
        ["手续费审批流",
         "申请人提交 → 财务审核 → 部门负责人审批 → 结束",
         "金额>10万增加总经理审批",
         "24小时未处理自动提醒",
         ""],
    ]

    for i, row_data in enumerate(flows):
        apply_data_row(ws, 4 + i, row_data, i)
        ws.row_dimensions[4 + i].height = 50

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 7: AI能力应用
# ══════════════════════════════════════════════════════════════
def create_sheet_07(wb):
    ws = wb.create_sheet("AI能力应用")
    max_col = 6
    apply_title(ws, "AI能力应用", max_col)

    headers = ["场景编号", "业务场景", "AI能力", "钉钉能力组件", "实现方式", "预期效果"]
    apply_header(ws, 3, headers)

    ai_data = [
        ["AI-01", "发票信息自动识别", "AI图片识别", "AI字段Agent-图片识别", "上传发票图片，AI自动提取发票号码、金额、日期", "减少80%手工录入"],
        ["AI-02", "智能数据填报", "AI生成文本", "AI字段Agent-AI生成文本", "输入供应商名称，AI自动补全信用代码、银行信息", "提升填报效率60%"],
        ["AI-03", "开票进度智能分析", "AI智能分析", "AI表格助理", '对话式查询"某供应商开票进度"', "零门槛数据分析"],
        ["AI-04", "经营数据可视化", "AI仪表盘", "AI仪表盘", "一句话生成经营看板", "10秒生成报表"],
        ["AI-05", "异常数据预警", "AI智能公式", "AI字段Agent-智能公式", "自动计算超期天数，标红预警", "风险提前3天发现"],
        ["AI-06", "合同条款审查", "AI文本分析", "AI字段Agent-AI生成文本", "AI辅助审查合同关键条款", "降低法务审查时间50%"],
        ["AI-07", "多语言供应商沟通", "AI翻译", "AI字段Agent-AI翻译", "供应商信息自动翻译13种语言", "支持跨国业务"],
        ["AI-08", "审批数据智能同步", "数据连接中心", "OA审批数据同步", "审批数据自动同步到AI表格", "零人工数据搬运"],
    ]

    for i, row_data in enumerate(ai_data):
        apply_data_row(ws, 4 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 8: 权限体系
# ══════════════════════════════════════════════════════════════
def create_sheet_08(wb):
    ws = wb.create_sheet("权限体系")
    max_col = 8
    apply_title(ws, "角色权限矩阵", max_col)

    headers = ["角色", "供应商数据", "需求数据", "合同数据", "付款数据", "开票数据", "手续费数据", "审批权限"]
    apply_header(ws, 3, headers)

    perm_data = [
        ["运营人员", "查看", "创建/查看本人", "查看", "查看", "查看/编辑", "创建/查看本人", "提交需求/手续费"],
        ["部门主管", "查看", "查看/审批本部门", "查看本部门", "查看本部门", "查看本部门", "查看/审批本部门", "审批需求/手续费"],
        ["财务人员", "查看/编辑", "查看全部", "创建/编辑/查看全部", "创建/编辑/查看全部", "编辑/查看全部", "查看/审批全部", "审批合同/付款/手续费"],
        ["法务人员", "查看", "查看", "查看/审批", "查看", "查看", "查看", "审批合同"],
        ["公章管理员", "查看", "查看", "查看", "查看", "查看", "查看", "审批用印"],
        ["财务总监", "查看/编辑全部", "查看全部", "查看/审批全部", "查看/审批全部", "查看全部", "查看/审批全部", "审批付款/手续费"],
        ["总经理", "查看/编辑全部", "查看全部", "查看/审批全部", "查看/审批全部", "查看全部", "查看/审批全部", "终审"],
        ["系统管理员", "全部权限", "全部权限", "全部权限", "全部权限", "全部权限", "全部权限", "流程配置"],
    ]

    for i, row_data in enumerate(perm_data):
        apply_data_row(ws, 4 + i, row_data, i)
        ws.cell(row=4 + i, column=1).font = BOLD_FONT

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 9: 报表看板设计
# ══════════════════════════════════════════════════════════════
def create_sheet_09(wb):
    ws = wb.create_sheet("报表看板设计")
    max_col = 8
    apply_title(ws, "报表看板设计", max_col)

    headers = ["报表编号", "报表名称", "报表类型", "数据源", "统计维度", "核心指标", "刷新频率", "钉钉能力"]
    apply_header(ws, 3, headers)

    report_data = [
        ["R-01", "经营总览看板", "仪表盘", "全部数据表", "时间/部门/供应商", "需求总额、已签约额、已开票额、差额、完成率", "实时", "AI仪表盘"],
        ["R-02", "需求趋势分析", "透视表", "需求申请表", "月度/季度", "需求笔数、金额趋势", "每日", "透视表视图"],
        ["R-03", "供应商开票进度", "看板视图", "开票进度表", "供应商/状态", "各供应商开票完成率", "实时", "看板视图"],
        ["R-04", "付款执行报表", "透视表", "付款申请表", "月度/供应商", "付款总额、待付款金额", "每日", "透视表视图"],
        ["R-05", "差额分析报表", "数据表", "需求+开票", "供应商/月份", "需求金额-已开票金额", "实时", "智能公式"],
        ["R-06", "手续费统计", "透视表", "手续费申请表", "月度/类型", "手续费总额、占比", "每日", "透视表视图"],
        ["R-07", "审批效率报表", "透视表", "OA审批同步数据", "流程类型/月份", "平均审批时长、超时率", "每周", "透视表视图"],
        ["R-08", "超期预警看板", "看板视图", "开票进度表", "超期天数", "超期笔数、最长超期天数", "实时", "看板+自动化"],
    ]

    for i, row_data in enumerate(report_data):
        apply_data_row(ws, 4 + i, row_data, i)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 10: 实施路径
# ══════════════════════════════════════════════════════════════
def create_sheet_10(wb):
    ws = wb.create_sheet("实施路径")
    max_col = 6
    apply_title(ws, "实施路径", max_col)

    headers = ["阶段", "阶段名称", "里程碑", "核心任务", "交付物", "验收标准"]
    apply_header(ws, 3, headers)

    impl_data = [
        ["P1", "基础搭建", "M1-数据模型上线",
         "创建7张数据表、配置字段与关联关系、导入供应商主数据",
         "数据模型文档、供应商数据表",
         "表结构完整、关联关系正确、主数据导入完成"],
        ["P2", "流程配置", "M2-审批流程上线",
         "配置5条审批流程、设置条件分支与超时策略、配置审批回调",
         "审批流程配置文档",
         "流程可正常发起与流转、条件分支正确、超时提醒生效"],
        ["P3", "自动化集成", "M3-数据自动流转",
         "配置OA审批数据同步、配置自动化工作流(8条)、配置消息通知",
         "自动化配置文档",
         "审批数据自动同步、工作流触发正确、通知及时送达"],
        ["P4", "AI能力部署", "M4-AI功能上线",
         "配置AI字段Agent(图片识别/文本生成)、配置AI仪表盘、配置透视表",
         "AI功能配置文档",
         "AI识别准确率>90%、仪表盘数据正确"],
        ["P5", "报表看板", "M5-报表体系上线",
         "创建8张报表/看板、配置权限与刷新频率",
         "报表清单与截图",
         "报表数据准确、刷新正常、权限正确"],
        ["P6", "测试上线", "M6-系统正式上线",
         "UAT测试、数据迁移、用户培训、上线切换",
         "测试报告、培训手册、上线报告",
         "全流程跑通、用户培训完成、无P0级缺陷"],
    ]

    for i, row_data in enumerate(impl_data):
        apply_data_row(ws, 4 + i, row_data, i)
        ws.row_dimensions[4 + i].height = 50

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 11: 风险评估
# ══════════════════════════════════════════════════════════════
def create_sheet_11(wb):
    ws = wb.create_sheet("风险评估")
    max_col = 8
    apply_title(ws, "风险评估", max_col)

    headers = ["风险编号", "风险类别", "风险描述", "发生概率", "影响程度", "风险等级", "应对策略", "责任人"]
    apply_header(ws, 3, headers)

    risk_data = [
        ["RK-01", "数据风险", "供应商主数据不完整或格式不统一", "高", "高", "高", "上线前进行数据清洗，制定数据录入规范", "项目经理"],
        ["RK-02", "流程风险", "审批流程节点设置不合理导致流程卡顿", "中", "高", "高", "上线前进行流程模拟测试，设置超时升级机制", "业务负责人"],
        ["RK-03", "采纳风险", "用户习惯改变导致系统使用率低", "中", "中", "中", "分阶段上线，提供培训与操作手册，设置过渡期", "项目经理"],
        ["RK-04", "技术风险", "AI字段Agent识别准确率不达预期", "低", "中", "低", "人工复核机制，持续优化AI提示词", "技术负责人"],
        ["RK-05", "容量风险", "数据量超出AI表格免费版限制", "中", "中", "中", "评估数据量，必要时升级企业版/旗舰版", "技术负责人"],
        ["RK-06", "集成风险", "OA审批与AI表格数据同步延迟", "低", "低", "低", "配置自动同步频率，设置手动同步兜底", "技术负责人"],
    ]

    for i, row_data in enumerate(risk_data):
        apply_data_row(ws, 4 + i, row_data, i)
        # 风险等级着色 (第6列)
        apply_risk_level(ws, 4 + i, 6)

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# Sheet 12: 成本估算
# ══════════════════════════════════════════════════════════════
def create_sheet_12(wb):
    ws = wb.create_sheet("成本估算")
    max_col = 5
    apply_title(ws, "成本估算", max_col)

    headers = ["费用项", "免费版", "企业版(18元/人/月)", "旗舰版(35元/人/月)", "说明"]
    apply_header(ws, 3, headers)

    cost_data = [
        ["单表最大行数", "20,000行", "30,000行(可扩容至千万)", "50,000行(可扩容至千万)", "根据数据量选择"],
        ["每月自动化运行次数", "500次", "50,000次", "500,000次", "根据业务频率选择"],
        ["高级权限自定义角色", "3个", "50个", "100个", "根据角色数量选择"],
        ["仪表盘联表计算", "不支持", "50个", "200个", "根据报表复杂度选择"],
        ["AI字段高速通道", "不支持", "支持", "支持", "影响AI处理速度"],
        ["OA审批高级版", "不含", "按需购买", "按需购买", "批量审批等高级API"],
        ["实施服务费", "-", "3-5万", "3-5万", "表单搭建+流程配置+培训"],
        ["年度运维费", "-", "1-2万/年", "1-2万/年", "系统维护+优化调整"],
    ]

    for i, row_data in enumerate(cost_data):
        apply_data_row(ws, 4 + i, row_data, i)
        ws.cell(row=4 + i, column=1).font = BOLD_FONT

    # 推荐方案
    r = 13
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1, value="推荐方案")
    cell.font = SUBTITLE_FONT
    cell.fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
    cell.border = THIN_BORDER

    r = 14
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    cell = ws.cell(row=r, column=1,
                   value="50人规模企业选择企业版，年费约10,800元+实施费4万≈5万首年投入")
    cell.font = Font(name="微软雅黑", bold=True, size=12, color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    cell.border = THIN_BORDER
    ws.row_dimensions[r].height = 36

    auto_column_width(ws)
    freeze_first_col(ws)


# ══════════════════════════════════════════════════════════════
# 主函数
# ══════════════════════════════════════════════════════════════
def main():
    wb = Workbook()

    create_sheet_01(wb)   # 项目概览
    create_sheet_02(wb)   # 业务流程分析
    create_sheet_03(wb)   # 解决方案架构
    create_sheet_04(wb)   # 功能模块清单
    create_sheet_05(wb)   # 数据模型设计
    create_sheet_06(wb)   # 审批流程设计
    create_sheet_07(wb)   # AI能力应用
    create_sheet_08(wb)   # 权限体系
    create_sheet_09(wb)   # 报表看板设计
    create_sheet_10(wb)   # 实施路径
    create_sheet_11(wb)   # 风险评估
    create_sheet_12(wb)   # 成本估算

    output_path = "钉钉AI表格+OA审批解决方案.xlsx"
    wb.save(output_path)
    print(f"[OK] Excel文件已生成: {output_path}")


if __name__ == "__main__":
    main()
